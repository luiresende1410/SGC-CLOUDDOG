import io
from typing import Any, Dict, List
import openpyxl

COLUNAS_OBRIGATORIAS = {"DATA", "NOME", "DEPARTAMENTO", "CARGO", "SALARIO", "TOTAL MENSAL"}

MAPEAMENTO_COLUNAS = {
    "DATA": "data",
    "NOME": "nome",
    "DEPARTAMENTO": "departamento",
    "CARGO": "cargo",
    "SALARIO": "salario",
    "BONUS (AWS)": "bonus_aws",
    "BONUS (PRD)": "bonus_prd",
    "COMISSOES": "comissoes",
    "HORA EXTRA": "hora_extra",
    "REMUNERACAO": "remuneracao",
    "REFEICAO": "refeicao",
    "TRANSPORTE": "transporte",
    "SEGURO SAUDE": "seguro_saude",
    "SEGURO VIDA": "seguro_vida",
    "FGTS": "fgts",
    "GPS": "gps",
    "EQUIPAMENTOS": "equipamentos",
    "ESCRITORIO": "escritorio",
    "FERIAS": "ferias",
    "13 SALARIO": "decimo_terceiro",
    "FGTS 2": "fgts_2",
    "GPS 2": "gps_2",
    "MULTA FGTS": "multa_fgts",
    "TOTAL MENSAL": "total_mensal",
    "SALARIO": "salario",
    "BONUS (AWS)": "bonus_aws",
    "BONUS (PRD)": "bonus_prd",
    "COMISSOES": "comissoes",
    "REMUNERACAO": "remuneracao",
    "REFEICAO": "refeicao",
    "SEGURO SAUDE": "seguro_saude",
    "FERIAS": "ferias",
    "13 SALARIO": "decimo_terceiro",
}

OBRIGATORIAS_CAMPOS = {"data", "nome", "departamento", "cargo", "salario", "total_mensal"}


def _normalizar(h: str) -> str:
    return h.strip().upper() if h else ""


def ler_planilha(arquivo_bytes: bytes) -> Dict[str, Any]:
    resultado: Dict[str, Any] = {"valido": False, "erros_criticos": [], "registros": []}
    try:
        wb = openpyxl.load_workbook(io.BytesIO(arquivo_bytes), data_only=True)
    except Exception as e:
        resultado["erros_criticos"].append(f"Erro ao abrir arquivo: {str(e)}")
        return resultado

    aba_nome = "Looker" if "Looker" in wb.sheetnames else wb.sheetnames[0]
    ws = wb[aba_nome]

    headers: List[str] = []
    header_row = None
    for row in ws.iter_rows():
        valores = [_normalizar(str(cell.value)) if cell.value is not None else "" for cell in row]
        if "NOME" in valores or "DATA" in valores:
            headers = valores
            header_row = row[0].row
            break

    if not headers:
        resultado["erros_criticos"].append("Cabecalho nao encontrado na planilha.")
        return resultado

    campos_presentes = set()
    for col in headers:
        if col:
            campo = MAPEAMENTO_COLUNAS.get(col, col.lower().replace(" ", "_"))
            campos_presentes.add(campo)

    faltando = OBRIGATORIAS_CAMPOS - campos_presentes
    if faltando:
        resultado["erros_criticos"].append(f"Colunas obrigatorias ausentes: {', '.join(sorted(faltando))}")
        return resultado

    resultado["valido"] = True

    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if not any(v is not None for v in row):
            continue
        registro: Dict[str, Any] = {}
        for i, header in enumerate(headers):
            if header and i < len(row):
                campo = MAPEAMENTO_COLUNAS.get(header, header.lower().replace(" ", "_"))
                registro[campo] = row[i]
        if registro.get("nome") and registro.get("data"):
            resultado["registros"].append(registro)

    return resultado
